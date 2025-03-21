import streamlit as st
import pandas as pd
import os
import numpy as np
from openpyxl.utils import get_column_letter
import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill
import importlib
import sys
from pathlib import Path

# 添加当前目录到 Python 路径
current_dir = Path(__file__).parent.absolute()
os.chdir(current_dir)  # 切换到脚本所在目录
sys.path.append(str(current_dir))
def adjust_column_width(worksheet):
    #设置第一行的背景色为黄色
    worksheet.row_dimensions[1].fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width

# 主页
def home():
    st.title("星云财务账单工具包")
    st.write("Create by Jackie Chen")
    st.write(f"Streamlit 版本: {st.__version__}")
    st.write(f"当前工作目录: {os.getcwd()}")
    st.write(f"当前时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


#main函数
def main():
    try:
        # 动态导入模块
        wdy_gcp = importlib.import_module('wdy_gcp').wdy_gcp
        xunlei_cdn_aws = importlib.import_module('xunlei_cdn_aws').xunlei_cdn_aws
        All_bill_microsoft = importlib.import_module('All_bill_microsoft').All_bill_microsoft
        tuo_bang_aws_9042 = importlib.import_module('tuo_bang_aws_9042').tuo_bang_aws_9042
        tuo_bang_aws_3677 = importlib.import_module('tuo_bang_aws_3677').tuo_bang_aws_3677
        xunlei_smartchain_aws = importlib.import_module('xunlei_smartchain_aws').xunlei_smartchain_aws
        zhenghao_aws = importlib.import_module('zhenghao_aws').zhenghao_aws

        st.sidebar.title('星云财务工具包')
        
        pages = {
            "主页": home,
            "万得云GCP翻译(ok)": wdy_gcp,
            "迅雷-CDN-AWS（ok）": xunlei_cdn_aws,
            "总账单-微软云(ok)": All_bill_microsoft,
            "拓邦-AWS-9042（ok）": tuo_bang_aws_9042,
            "拓邦-AWS-3677": tuo_bang_aws_3677, 
            "迅雷-Smartchain-AWS（ok）": xunlei_smartchain_aws,
            "正浩-AWS(ok)": zhenghao_aws
        }
        
        page = st.sidebar.selectbox("选择功能页面", list(pages.keys()))
        pages[page]()
    except Exception as e:
        st.error(f"加载模块时出错: {str(e)}")
        st.error(f"当前工作目录: {os.getcwd()}")
        st.error(f"Python 路径: {sys.path}")
        st.error(f"可用文件: {os.listdir('.')}")

if __name__ == "__main__":
    main()



